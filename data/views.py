from django.shortcuts import render
from django.db.models import Q, Count
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view
from rest_framework.response import Response
import re
import csv
from django.http import HttpResponse
import pandas as pd

from .models import SensorHeartRate, Student, Session, Activity
from .serializers import SensorHeartRateSerializer, StudentSerializer, SessionSerializer, ActivitySerializer, \
    StudentActivitiesSerializer, SensorHeartRateActivitiesSerializer
from .utils.date_converter import EpochMsDateTimeField
from .utils.hrv import compute_metrics_from_ibi_list_manual, compute_metrics_from_ibi_list_lib, compute_metrics_from_hr_list, \
    compute_metrics_from_deprecated_ibi_list
from .utils.parsers import parse_bool, excel_text, column_exists

from openpyxl import Workbook


#<editor-fold desc="Get All - for test">

# --------- Get All - for test ---------

@api_view(['GET'])
def sensor_heart_rate(request):
    if request.method == 'GET':
        data = SensorHeartRate.objects.all()[:30]
        sensor_heart_rate_serializer = SensorHeartRateSerializer(data, many=True)
        return Response(sensor_heart_rate_serializer.data)

@api_view(['GET'])
def student(request):
    if request.method == 'GET':
        data = Student.objects.all()[:30]
        student_serializer = StudentSerializer(data, many=True)
        return Response(student_serializer.data)

@api_view(['GET'])
def session(request):
    if request.method == 'GET':
        data = Session.objects.all()[:30]
        session_serializer = SessionSerializer(data, many=True)
        return Response(session_serializer.data)

@api_view(['GET'])
def activity(request):
    if request.method == 'GET':
        data = Activity.objects.all()[:30]
        activity_serializer = ActivitySerializer(data, many=True)
        return Response(activity_serializer.data)

#</editor-fold>

#<editor-fold desc="Index page">

# --------- Index page ---------
def data_index(request):
    return render(request, "data_index.html")

#</editor-fold>

#<editor-fold desc="Student - Activities page">

# --------- Student - Activities page ---------

def student_activities_page(request):

    student_id_raw = (request.GET.get("student_id") or "").strip()
    done_raw = (request.GET.get("done") or "").strip()
    types_raw = (request.GET.get("types") or "").strip()
    limit_raw = (request.GET.get("limit") or "-1").strip()
    ibi_count_raw = (request.GET.get("ibi_count") or "120").strip()

    # parse limit
    try:
        limit = max(-1, min(int(limit_raw), 200))
    except ValueError:
        limit = -1

    # parse ibi count
    try:
        ibi_count = max(-1, int(ibi_count_raw))
    except ValueError:
        ibi_count = -1

    # parse done flag
    done_val = parse_bool(done_raw)

    done_select = ""
    if done_val is True:
        done_select = "1"
    elif done_val is False:
        done_select = "0"

    # parse activity types
    types = [t.strip() for t in types_raw.split(",") if t.strip()]
    type_q = Q()
    for t in types:
        type_q |= Q(activityType__icontains=t)

    results = []
    matched_session_ids = []
    matched_activities_ids = []
    student_nickname = None
    error = None

    students = Student.objects.only("id", "nickname").order_by("id")

    if not student_id_raw:
        pass
    elif not student_id_raw.isdigit():
        error = "Provide a numeric student_id."
    else:
        student_id = int(student_id_raw)

        # sessions for student
        sess_qs = Session.objects.filter(student_id=student_id).order_by("id")
        matched_session_ids = list(sess_qs.values_list("id", flat=True))

        # activities
        act_qs = Activity.objects.select_related("session", "session__student").filter(session_id__in=matched_session_ids)

        has_performance_column = column_exists("activity", "performance")
        if not has_performance_column:
            act_qs = act_qs.values("id", "activityType", "session_id", "session__student_id", "done", "level")
        else:
            act_qs = act_qs.values("id", "activityType", "session_id", "session__student_id", "done", "level", "performance")

        # done
        if done_val is not None:
            act_qs = act_qs.filter(done=done_val)

        # activityType
        if types:
            act_qs = act_qs.filter(type_q)

        act_qs = act_qs.annotate(row_count=Count("heart_rates"))
        if limit != -1:
            act_qs = act_qs.order_by("session_id", "id")[:limit]
        else:
            act_qs = act_qs.order_by("session_id", "id")

        results = StudentActivitiesSerializer(act_qs, many=True).data
        #results = StudentActivitiesSerializer(list(act_qs), many=True).data

        matched_activities_ids = list(act_qs.values_list("id", flat=True))

        # add valid_ibi_count for each activity
        filtered_results = []
        for r in results:
            act_id = r["activity_id"] if "activity_id" in r else r["id"]
            hr_qs = SensorHeartRate.objects.filter(activity_id=act_id).only(
                "value_ibi_0", "status_ibi_0",
                "value_ibi_1", "status_ibi_1",
                "value_ibi_2", "status_ibi_2",
                "value_ibi_3", "status_ibi_3"
            )
            valid_count = 0
            for hr in hr_qs:
                for i in range(4):
                    val = getattr(hr, f"value_ibi_{i}")
                    st = getattr(hr, f"status_ibi_{i}")
                    if st == 110 and val is not None:
                        valid_count += 1

            r["valid_ibi_count"] = valid_count

            # apply filter only if ibi_count >= 0
            if ibi_count == -1 or valid_count >= ibi_count:
                filtered_results.append(r)

        results = filtered_results

        # recompute matched activity and session ids after ibi_count filter
        matched_activities_ids = [r["activity_id"] if "activity_id" in r else r["id"] for r in results]
        matched_session_ids = list(
            Activity.objects.filter(id__in=matched_activities_ids)
            .values_list("session_id", flat=True)
            .distinct()
        )

        # student nickname
        try:
            student_obj = Student.objects.get(pk=student_id)
            student_nickname = student_obj.nickname or None
        except Student.DoesNotExist:
            pass

    context = {
        "params": {
            "student_id": student_id_raw,
            "done": done_raw,
            "done_select": done_select,
            "types": types_raw,
            "limit": limit,
            "ibi_count": ibi_count,
        },
        "students": students,
        "student_nickname": student_nickname,
        "matched_session_ids": matched_session_ids,
        "matched_activities_ids": matched_activities_ids,
        "results": results,
        "count": len(results),
        "error": error,
    }
    return render(request, "student_activities.html", context)

#</editor-fold>

#<editor-fold desc="Sensor Heart Rate - Activities page">

# --------- Sensor Heart Rate - Activities page ---------

def sensor_heart_rate_activities_page(request):

    activity_ids_raw = (request.GET.get("activity_ids") or "").strip()
    limit_raw = (request.GET.get("limit") or "-1").strip()

    # parse limit
    try:
        limit = max(-1, min(int(limit_raw), 500))
    except ValueError:
        limit = -1

    # extract activities ids
    activity_ids = [int(x) for x in re.findall(r"\d+", activity_ids_raw)]
    activity_ids = list(dict.fromkeys(activity_ids))

    matched_session_ids = []
    results = []
    count = 0
    qs = None

    if activity_ids:

        matched_session_ids = sorted(set(
            Activity.objects.filter(id__in=activity_ids)
            .values_list("session_id", flat=True)
        ))

        if limit != -1:
            qs = (SensorHeartRate.objects
                .filter(activity_id__in=activity_ids)
                .select_related("activity", "session", "student")
                .order_by("activity_id", "id")[:limit])
        else:
            qs = (SensorHeartRate.objects
                .filter(activity_id__in=activity_ids)
                .select_related("activity", "session", "student")
                .order_by("activity_id", "id"))

        #has_performance_column = column_exists("activity", "performance")
        #if not has_performance_column:
        qs = qs.values("id", "activity_id", "session_id", "student_id", "value_heart_rate", "status_heart_rate",
                       "value_ibi_0", "status_ibi_0", "value_ibi_1", "status_ibi_1", "value_ibi_2", "status_ibi_2",
                       "value_ibi_3", "status_ibi_3", "value_ibi_depr", "status_ibi_depr", "timestamp")

        results = SensorHeartRateActivitiesSerializer(qs, many=True).data
        count = len(results)

    # flags download the full CSV and the processed CSV
    download = (request.GET.get("download") or "").strip().lower()
    download_processed = (request.GET.get("download_processed") or "").strip().lower()

    # regular CSV download
    if download == "1":
        headers = [
            "id",
            "activity_id", "session_id", "student_id",
            "value_heart_rate", "status_heart_rate",
            "value_ibi_0", "status_ibi_0",
            "value_ibi_1", "status_ibi_1",
            "value_ibi_2", "status_ibi_2",
            "value_ibi_3", "status_ibi_3",
            "value_ibi_depr", "status_ibi_depr",
            "timestamp"
        ]
        resp = HttpResponse(content_type="text/csv")
        fname = f"sensor_heart_rate__act_{activity_ids_raw}.csv"
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        w = csv.writer(resp)
        w.writerow(headers)
        for row in results:
            out = []
            for h in headers:
                cell = row.get(h, "")
                if h == "timestamp":
                    cell = excel_text(cell)
                out.append(cell)
            w.writerow(out)
        return resp

    # processed CSV download
    if download_processed == "1":
        fname = f"processed_sensor_heart_rate__act_{activity_ids_raw}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'

        wb = Workbook()

        # --- Sheet 1: IBI values ---
        ws1 = wb.active
        ws1.title = "IBI"
        ws1.append(["id", "activity_id", "session_id", "student_id", "value_ibi", "timestamp"])

        fmt_ts = EpochMsDateTimeField().to_representation
        if qs is None:
            qs = SensorHeartRate.objects.none()

        for rec in qs:
            # base = [rec.id, rec.activity_id, rec.session_id, rec.student_id]
            base = [rec["id"], rec["activity_id"], rec["session_id"], rec["student_id"]]
            ts_str = fmt_ts(rec.get("timestamp"))
            for i in range(4):
                val = rec.get(f"value_ibi_{i}")
                st = rec.get(f"status_ibi_{i}")
                if st == 110 and val is not None:
                    ws1.append(base + [val, ts_str])

        # --- Sheet 2: IBI deprecated values ---
        ws2 = wb.create_sheet("IBI_Depr")
        ws2.append(["id", "activity_id", "session_id", "student_id", "value_ibi_depr", "timestamp"])

        for rec in qs:
            if rec.get("status_ibi_depr") == 110 and rec.get("value_ibi_depr") is not None:
                ws2.append([
                    rec["id"], rec["activity_id"], rec["session_id"], rec["student_id"],
                    rec["value_ibi_depr"],
                    fmt_ts(rec.get("timestamp"))
                ])

        # --- Sheet 3: Heart rate values ---
        ws3 = wb.create_sheet("HeartRate")
        ws3.append(["id", "activity_id", "session_id", "student_id", "value_heart_rate", "timestamp"])

        for rec in qs:
            if rec.get("status_heart_rate") == 10 and rec.get("value_heart_rate") is not None:
                ws3.append([
                    rec["id"], rec["activity_id"], rec["session_id"], rec["student_id"],
                    rec["value_heart_rate"],
                    fmt_ts(rec.get("timestamp"))
                ])

        wb.save(resp)
        return resp

    context = {
        "params": {"activity_ids": activity_ids_raw, "limit": limit},
        "matched_activity_ids": activity_ids,
        "matched_session_ids": matched_session_ids,
        "results": results,
        "count": count,
    }
    return render(request, "sensor_heart_rate_activities.html", context)

#</editor-fold>

#<editor-fold desc="HRV-Metrics page">

# --------- HRV-Metrics page ---------

def sliding_windows(data, window_size=120, step=30):
    n = len(data)
    if n == 0:
        return

    start = 0
    while start + window_size <= n:
        end = start + window_size
        yield start, end, data[start:end]
        start += step

    # Handle remaining tail
    if start < n:
        end = n
        start = max(0, n - window_size)
        yield start, end, data[start:end]

def compute_windowed_metrics(act_id, act_type, ibi_list, id_list, hr_values=None, window_size=120, step=30):
    rows = []

    has_performance = column_exists("activity", "performance")
    performance_value = None
    if has_performance:
        try:
            performance_value = (Activity.objects.filter(id=act_id).values_list("performance", flat=True).first())
        except Exception as e:
            performance_value = None

    for start, end, ibi_window in sliding_windows(ibi_list, window_size=window_size, step=step):
        id_window = id_list[start:end]
        hr_window = hr_values[start:end] if hr_values else []

        metrics = compute_metrics_from_ibi_list_lib(ibi_window, id_window)
        metrics.update(compute_metrics_from_hr_list(hr_window, ibi_window))

        row = {
            "activity_id": act_id,
            "activityType": act_type,
            "window_start": start + 1,
            "window_end": end,
            "performance": performance_value if has_performance else None,
            **{k: (None if pd.isna(v) else float(v)) for k, v in metrics.items()},
        }
        rows.append(row)

    return rows

@require_http_methods(["GET", "POST"])
def hrv_metrics_page(request):

    results = []
    results_lib = []
    results_lib_windows = []
    results_depr = []
    error = None

    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]
        filename = f.name.lower()

        try:
            if filename.endswith(".csv"):
                df = pd.read_csv(f)
                df_hr = None
                df_depr = None
            elif filename.endswith((".xls", ".xlsx")):
                dfs = pd.read_excel(f, sheet_name=None)
                df = dfs.get("IBI")
                df_hr = dfs.get("HeartRate")
                df_depr = dfs.get("IBI_Depr")
            else:
                raise ValueError("Unsupported file format. Upload CSV or XLSX.")
        except Exception as e:
            error = f"Failed to read file: {e}"
            return render(request, "hrv_metrics.html", { "error": error })

        required_cols = {"id", "activity_id", "session_id", "student_id", "value_ibi"}
        if df is None or not required_cols.issubset(set(df.columns)):
            missing = required_cols - set(df.columns)
            error = f"Missing required columns: {', '.join(missing)}"
            return render(request, "hrv_metrics.html", { "error": error })

        # ensure numeric types
        df["activity_id"] = pd.to_numeric(df["activity_id"], errors="coerce")
        df["session_id"] = pd.to_numeric(df["session_id"], errors="coerce")
        df["student_id"] = pd.to_numeric(df["student_id"], errors="coerce")
        df["value_ibi"] = pd.to_numeric(df["value_ibi"], errors="coerce")

        # Prepare HR data if available
        if df_hr is not None and "value_heart_rate" in df_hr.columns:
            df_hr["activity_id"] = pd.to_numeric(df_hr["activity_id"], errors="coerce")
            df_hr["value_heart_rate"] = pd.to_numeric(df_hr["value_heart_rate"], errors="coerce")

        # Prepare IBI deprecated data if available
        if df_depr is not None and "value_ibi_depr" in df_depr.columns:
            df_depr["activity_id"] = pd.to_numeric(df_depr["activity_id"], errors="coerce")
            df_depr["value_ibi_depr"] = pd.to_numeric(df_depr["value_ibi_depr"], errors="coerce")

        act_ids = df["activity_id"].dropna().astype(int).unique().tolist()
        act_types = dict(Activity.objects.filter(id__in=act_ids).values_list("id", "activityType"))

        # parse checkboxes
        include_lib = request.POST.get("include_lib") == "1"
        include_custom = request.POST.get("include_custom") == "1"
        include_window = request.POST.get("include_window") == "1"
        include_deprecated = request.POST.get("include_deprecated") == "1"

        has_performance = column_exists("activity", "performance")

        ## compute deprecated IBI
        if df_depr is not None and include_deprecated:
            for act_id, g in df_depr.groupby("activity_id", dropna=True):
                ibi_depr_ms = g["value_ibi_depr"].dropna().astype(float).values.tolist()
                id_depr_list = g["id"].dropna().astype(float).values.tolist()
                act_id_int = int(act_id) if pd.notna(act_id) else None

                hr_values = []
                if df_hr is not None:
                    hr_values = df_hr[df_hr["activity_id"] == act_id_int]["value_heart_rate"].dropna().astype(float).tolist()

                ibi_ms = []
                if df is not None:
                    ibi_ms = df[df["activity_id"] == act_id_int]["value_ibi"].dropna().astype(float).tolist()

                metrics_depr = compute_metrics_from_deprecated_ibi_list(ibi_depr_ms, id_depr_list)
                metrics_depr.update(compute_metrics_from_hr_list(hr_values, ibi_ms))

                performance_value = None
                if has_performance:
                    try:
                        performance_value = (Activity.objects.filter(id=act_id).values_list("performance", flat=True).first())
                    except Exception as e:
                        performance_value = None

                row_depr = {
                    "activity_id": act_id_int,
                    "activityType": act_types.get(act_id_int),
                    "performance": performance_value if has_performance else None,
                    **{k: (None if pd.isna(v) else float(v)) for k, v in metrics_depr.items()},
                }
                results_depr.append(row_depr)

        #### main computing
        for act_id, g in df.groupby("activity_id", dropna=True):
            ibi_ms = g["value_ibi"].dropna().astype(float).values.tolist()
            id_list = g["id"].dropna().astype(float).values.tolist()
            act_id_int = int(act_id) if pd.notna(act_id) else None

            hr_values = []
            if df_hr is not None:
                hr_values = df_hr[df_hr["activity_id"] == act_id_int]["value_heart_rate"].dropna().astype(float).tolist()

            performance_value = None
            if has_performance:
                try:
                    performance_value = (Activity.objects.filter(id=act_id).values_list("performance", flat=True).first())
                except Exception as e:
                    performance_value = None

            ## compute custom / manual HRV
            if include_custom:
                metrics = compute_metrics_from_ibi_list_manual(ibi_ms)
                metrics.update(compute_metrics_from_hr_list(hr_values, ibi_ms))
                row = {
                    "activity_id": act_id_int,
                    "activityType": act_types.get(act_id_int),
                    "performance": performance_value if has_performance else None,
                    **{k: (None if pd.isna(v) else float(v)) for k, v in metrics.items()},
                }
                results.append(row)

            ## compute HRV using PyHRV
            if include_lib:
                metrics_lib = compute_metrics_from_ibi_list_lib(ibi_ms, id_list)
                metrics_lib.update(compute_metrics_from_hr_list(hr_values, ibi_ms))
                row_lib = {
                    "activity_id": act_id_int,
                    "activityType": act_types.get(act_id_int),
                    "performance": performance_value if has_performance else None,
                    **{k: (None if pd.isna(v) else float(v)) for k, v in metrics_lib.items()},
                }
                results_lib.append(row_lib)

            # compute sliding windowed metrics
            if include_window:
                windowed_rows = compute_windowed_metrics(act_id_int, act_types.get(act_id_int), ibi_ms, id_list, hr_values, window_size=120, step=30)
                results_lib_windows.extend(windowed_rows)

        # sort by activity_id
        if include_custom:
            results.sort(key=lambda r: (r["activity_id"] is None, r["activity_id"]))
        if include_lib:
            results_lib.sort(key=lambda r: (r["activity_id"] is None, r["activity_id"]))
        if include_window:
            results_lib_windows.sort(key=lambda r: (r["activity_id"] is None, r["activity_id"]))
        if include_deprecated:
            results_depr.sort(key=lambda r: (r["activity_id"] is None, r["activity_id"]))

        # CSV download
        if (request.POST.get("download") or "").lower() == "1":
            headers = ["activity_id", "activityType", "performance",
                       "mean_hr", "mean_rr",
                       "rmssd", "rmssd_chunks", "sdnn", "sdnn_chunks",
                       "nn50", "nn50_chunks", "pnn50", "pnn50_chunks", "tinn", "tinn_chunks",
                       "stress_index", "pns_index", "sns_index",
                       "lf", "hf", "lf_hf",
                       "sd1", "sd2", "sd2_sd1",
                       "ap_en", "samp_en", "dfa_a1", "dfa_a2"]
            headers_2_decimals = [ "performance", "mean_hr", "mean_rr", "stress_index" ]
            resp = HttpResponse(content_type="text/csv")
            fname = f"hrv_metrics.csv"
            resp["Content-Disposition"] = f'attachment; filename="{fname}"'
            w = csv.writer(resp)
            w.writerow(headers)

            ## download the first checked list
            results_to_download = []
            if include_window:
                results_to_download = results_lib_windows
            elif include_lib:
                results_to_download = results_lib
            elif include_custom:
                results_to_download = results
            elif include_deprecated:
                results_to_download = results_depr

            for r in results_to_download:
                row = []
                for h in headers:
                    val = r.get(h, "")
                    if isinstance(val, float):
                        if h in headers_2_decimals:
                            row.append(f"{val:.2f}")
                        else:
                            row.append(f"{val:.3f}")
                    else:
                        row.append(val)
                w.writerow(row)
            return resp

    return render(request, "hrv_metrics.html", {"results": results,
                                                                   "results_lib": results_lib,
                                                                   "results_lib_windows": results_lib_windows,
                                                                   "results_depr": results_depr,
                                                                   "error": error})

#</editor-fold>

#<editor-fold desc="HRV-Interpretation page">

# --------- HRV-Interpretation page ---------

@require_http_methods(["GET", "POST"])
def hrv_interpretation_page(request):

    results = []
    error = None
    relax_means = {}

    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]

        try:
            df = pd.read_csv(f)
        except Exception as e:
            error = f"Failed to read CSV: {e}"
            return render(request, "hrv_interpretation.html", {"error": error})

        required_cols = {"activity_id", "activityType",
                         "mean_hr", "mean_rr",
                         "rmssd", "sdnn", "nn50", "pnn50", "tinn",
                         "rmssd_chunks", "sdnn_chunks", "nn50_chunks", "pnn50_chunks", "tinn_chunks",
                         "stress_index", "pns_index", "sns_index", "lf", "hf", "lf_hf",
                         "sd1", "sd2", "sd2_sd1", "ap_en", "samp_en", "dfa_a1", "dfa_a2"}
        if column_exists("activity", "performance"):
            required_cols.add("performance")

        if not required_cols.issubset(set(df.columns)):
            missing = required_cols - set(df.columns)
            error = f"Missing required columns: {', '.join(missing)}"
            return render(request, "hrv_interpretation.html", {"error": error})

        # Compute relax (delaygratification) average
        relax_df = df[df["activityType"] == "delaygratification"]
        if not relax_df.empty:
            relax_means = relax_df.mean(numeric_only=True).to_dict()

        # check if analyze is requested
        analyzed = (request.POST.get("analyze") or "").lower() == "1"

        # expected directions
        decrease_metrics = ["mean_rr",
                            "rmssd", "rmssd_chunks", "sdnn", "sdnn_chunks",
                            "nn50", "nn50_chunks", "pnn50", "pnn50_chunks", "tinn", "tinn_chunks",
                            "lf", "hf", "sd1", "sd2", "sd2_sd1",
                            "ap_en", "samp_en", "dfa_a1", "dfa_a2", "pns_index"]
        increase_metrics = ["mean_hr", "lf_hf", "stress_index", "sns_index"]

        for _, row in df.iterrows():
            r = {col: row[col] for col in required_cols}
            r["styles"] = {}

            if analyzed and relax_means and row["activityType"] != "delaygratification":
                for m in decrease_metrics:
                    val = r[m]
                    baseline = relax_means.get(m)
                    if pd.notna(val) and pd.notna(baseline) and val != 0:
                        r["styles"][m] = "background-color: #c6efce;" if val < baseline else "background-color: #ffc7ce;"

                for m in increase_metrics:
                    val = r[m]
                    baseline = relax_means.get(m)
                    if pd.notna(val) and pd.notna(baseline) and val != 0:
                        r["styles"][m] = "background-color: #c6efce;" if val > baseline else "background-color: #ffc7ce;"

            results.append(r)

    return render(request, "hrv_interpretation.html", {"results": results, "error": error})

#</editor-fold>

#<editor-fold desc="CSV Merging page">

# --------- CSV Merging page ---------

def merge_csv_page(request):
    error = None
    merged_df = None

    if request.method == "POST" and request.FILES.getlist("files"):
        files = request.FILES.getlist("files")
        dfs = []
        for f in files:
            try:
                df = pd.read_csv(f)
                dfs.append(df)
            except Exception as e:
                error = f"Failed to read {f.name}: {e}"
                return render(request, "merge_csv.html", {"error": error})

        if dfs:
            merged_df = pd.concat(dfs, ignore_index=True)
            resp = HttpResponse(content_type="text/csv")
            resp["Content-Disposition"] = 'attachment; filename="merged_hrv_metrics.csv"'
            merged_df.to_csv(resp, index=False)
            return resp

    return render(request, "merge_csv.html", {"error": error})

#</editor-fold>

#<editor-fold desc="Performance Analysis page">

# --------- Performance Analysis page ---------

def performance_analysis_page(request):

    return render(request, "performance_analysis.html", {"error": ""})

#</editor-fold>